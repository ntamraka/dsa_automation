import os
import argparse
from argparse import ArgumentParser
import time
import datetime
import openpyxl
from pathlib import Path
import pandas as pd
import numpy as np


VERSION=1.0

class Parser():

    def __init__(self, args):
        self.args = args
        self.output_dir=self.args.output_dir

    def Parsing(self):
        counter_pd=[]
        counters = ["filename","metric_CPU utilization %","metric_CPU operating frequency (in GHz)",
        "metric_DRAM power (watts)","metric_package power (watts)"]
        filelist = os.listdir(self.output_dir)
        print(counters)
        for file in filelist:
            xlsx_file = Path(self.output_dir, file)
            wb_obj = openpyxl.load_workbook(xlsx_file)
            sheet = wb_obj.active

            counters_list=[]
            counters_list.append(file)
            for row in sheet.iter_rows(1, sheet.max_row):
              for counter in counters:
                if row[0].value==counter:
                    counters_list.append(row[1].value)
            
            counter_pd.append(counters_list)
            print(counters_list)
            #print(sheet["A176"].value,",",sheet["B176"].value)  
            #print(sheet["A177"].value," ",sheet["B177"].value) 
            #print(sheet["A177"].value," ",sheet["B177"].value) 
        df = pd.DataFrame(counter_pd, columns = counters)
        print(df)
        df.to_csv(self.output_dir+'/summary.csv')
            

if __name__ == "__main__":
    print("Using version :" ,VERSION)
    parser = ArgumentParser()
    parser.add_argument("-p", "--output_dir", help="Specify Path of logs",default=".")
    args = parser.parse_args()
    Automation=Parser(args)
    Automation.Parsing()


