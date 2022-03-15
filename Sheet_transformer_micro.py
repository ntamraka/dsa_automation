import openpyxl as xl;
import pandas as pd
import argparse
from argparse import ArgumentParser
parser = ArgumentParser()
parser.add_argument("-p", "--output_dir", help="Specify Path of logs",default="./Automation_log/")
parser.add_argument("-o", "--output", help="Specify name of summary",default="summary.csv")
args = parser.parse_args()
read_file = pd.read_csv (args.output_dir+'summary.csv')
read_file.to_excel (args.output_dir+'summary.xlsx', index = None, header=False) 

wb1 = xl.load_workbook(args.output_dir+'summary.xlsx') 
#print(wb1)
ws1 = wb1.worksheets[0] 
wb = xl.Workbook()
sheet_obj = wb.active

def filldata(word,j):
    for k in range(len(word)):
        obj = sheet_obj.cell(row = j,column = k+1)
        obj.value = word[k]

list_of_op=[4,3,6,9,16]
list_of_queuedepth=[1,2,4,8,16,32,64,128]
#list_of_ts=["qd/ts",1024,2048,4096,8192,16384,65536,131072,262144]
list_of_ts=["qd/ts",'1K','2K','4K','8K','16K','32K','64K','128K']
mr = ws1.max_row 
mc = ws1.max_column 
row_count=1
for op in list_of_op:
    filldata([op],row_count)
    row_count=row_count+1
    filldata(list_of_ts,row_count)
    for queuedepth in list_of_queuedepth:
        row_count=row_count+1
        word = [] 
        word.append(queuedepth)
        for i in range (1, mr + 1):
            BW=(ws1.cell(row = i, column=9).value)
            #print(BW)
            op_1=(ws1.cell(row = i, column=1).value)
            qd_1=(ws1.cell(row = i, column=2).value)
            if (op_1==op and queuedepth==int(qd_1)):
                word.append(ws1.cell(row = i, column=9).value)     
            filldata(word,row_count)   
    row_count=row_count+1       
wb.save(str(args.output_dir+"single_instance.xlsx"))
